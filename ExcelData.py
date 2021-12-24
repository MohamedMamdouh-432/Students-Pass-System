from openpyxl import load_workbook
from tkinter import messagebox as msgx

wb = load_workbook('CollegeData.xlsx')
ws = wb.active

def AddExcel(which, data):
    if which == 'std' :
        ws = wb['Students']
    elif which == 'lec' :
        ws = wb['Lecturers']
    ws.append(data)
    wb.save('CollegeData.xlsx')

def SearchExcel(which, id):
    if which == 'std' :
        ws = wb['Students']
    elif which == 'lec' :
        ws = wb['Lecturers']
    for r in range(2,ws.max_row+1):
        com_id = str(ws.cell(row=r, column=3).value)
        if com_id == id :
            return r
    return 0

def CheckExcel(which, id):
    r = SearchExcel(which, id)
    if r != 0 :
        msgx.showinfo('Check Manager', 'Person is already Existed')
    else :
        msgx.showinfo('Check Manager', 'Person is not Existed')

def DeleteExcel(which, id):
    r = SearchExcel(which, id)
    if r != 0 :
        if which == 'std' :
            ws = wb['Students']
        elif which == 'lec' :
            ws = wb['Lecturers']
        ws.delete_rows(r)
        msgx.showinfo('Delete Manager', 'Person is already Deleted')
    else:
        msgx.showinfo('Delete Manager', 'Person is not Existed')
    wb.save('CollegeData.xlsx')

def ShowExcel(which, id):
    r = SearchExcel(which, id)
    if r != 0 :
        msgx.showinfo('Check Manager', 'Person is already Existed')
        if which == 'std' :
            ws = wb['Students']
            data = {
                'Name' : ws.cell(row=r, column=1).value ,
                'National Id' : ws.cell(row=r, column=2).value ,
                'College Id' : ws.cell(row=r, column=3).value ,
                'Departement' : ws.cell(row=r, column=4).value ,
                'Vaccination State' : ws.cell(row=r, column= 5).value ,
                'Current Year' : ws.cell(row=r, column=6).value ,
                'Graduation Year' : ws.cell(row=r, column=7).value ,
                'Category' : ws.cell(row=r, column=8).value
            }
        elif which == 'lec' :
            ws = wb['Lecturers']
            data = {
                'Name' : ws.cell(row=r, column=1).value ,
                'National Id' : ws.cell(row=r, column=2).value ,
                'College Id' : ws.cell(row=r, column=3).value ,
                'Departement' : ws.cell(row=r, column=4).value ,
                'Vaccination State' : ws.cell(row=r, column= 5).value ,
                'Speciality' : ws.cell(row=r, column=6).value ,
                'Source University' : ws.cell(row=r, column=7).value
            }
        return data
    else :
        msgx.showinfo('Check Manager', 'Person is not Existed')
        return 0