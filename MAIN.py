from tkinter import *
from System import CollegeSystem
import os

if __name__ == "__main__":
    if os.path.exists('CollegeData.xlsx') == False:
        from openpyxl import Workbook
        wb = Workbook()
        wb.create_sheet('Students')
        wb.create_sheet('Lecturers')
        wb.remove('Sheet')
        data = ['Name', 'Nationa4l Id', 'College Id', 'Departement', 'Vaccination State', 'Current Year', 'Graduation Year', 'Category']
        wb['Students'].append(data)
        data = ['Name', 'National Id', 'College Id', 'Departement', 'Vaccination State', 'Speciality', 'Source University']
        wb['Lecturers'].append(data)
        wb.save('CollegeData.xlsx')

    SystemFrame = Tk()
    SystemFrame.geometry("800x700")
    SystemFrame.configure(background="#1090b0")
    SystemFrame.title("College Pass System")
    SystemFrame.iconbitmap('p3.ico')
    System = CollegeSystem(SystemFrame)
    SystemFrame.mainloop()