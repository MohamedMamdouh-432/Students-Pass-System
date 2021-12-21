from tkinter import *
from openpyxl import Workbook
from System import CollegeSystem

if __name__ == "__main__":
    wb = Workbook()
    wb.create_sheet("Students", 0)
    wb.create_sheet("Lecturers", 1)
    wb.save('CollegeData.xlsx')

    SystemFrame = Tk()
    SystemFrame.geometry("700x400")
    SystemFrame.configure(background="light green")
    System = CollegeSystem(SystemFrame)
    SystemFrame.mainloop()